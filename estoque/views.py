from .funcoes_comunidades import *
from .funcoes_produtos import *
from .funcoes_vendas import *
from usuarios.funcoes_usuarios import *
from usuarios.funcoes_familias import *
from django.shortcuts import render
from .forms import ProdutoForm, ComunidadeForm
from .models import Produto, NomeProduto, Vendas, LogsItens, P_Excel, VendasControle
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.core.files.uploadedfile import InMemoryUploadedFile
from rolepermissions.decorators import has_permission_decorator
from django.shortcuts import get_object_or_404
from django.contrib import auth
from django.template.defaultfilters import slugify
from django.http import JsonResponse
import os
import re
import json
import itertools
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.filters import AutoFilter
from django.db.models import Q, ProtectedError, Max
from django.utils import timezone
from django.core.paginator import Paginator
from django.dispatch import Signal, receiver
from usuarios.models import Users
from estoque.signals import produto_deleted, novonome_produto_deleted, vendas_deleted, vendas_geral_deleted
from urllib.parse import urlencode
from django.db import transaction
import unidecode
from decimal import Decimal, ROUND_DOWN, ROUND_UP

#Busca = Q(
#        Q(nome_produto_id='2') & | Q(tamanho_produto_id=='2')     
# ) 
# python -m pip install --upgrade --force-reinstall pip and then do python -m pip freeze (Comando pra reinstalar o python corretamente ao trocar o diretório)
#Venda_finalizada = 0 "ainda não foi finalizada"
#Venda_finalizada = 1 "foi finalizada"
#Venda_finalizada = 2 "estorno"
# #Não apagar


#Função para a tela de logs
@has_permission_decorator('acessar_logs')#Apenas o ADMIN e o Padre tem acesso.
def listar_logs(request):
    if request.method == "GET":
        nome_user = request.GET.get('nome_user')
        dia = request.GET.get('dia')
        model = request.GET.get('model')
        acao = request.GET.get('acao')
        paginacao = LogsItens.objects.all()
        
        validacao, page = Get_Paginacao_Logs(request, nome_user, dia, acao, model, paginacao)
        if validacao:
            return validacao
        url_atual = request.path
        context = {
            'page': page,
            'url_atual': url_atual,
        }
        return render(request, 'listar_logs.html', context)


#Função de redirect para tela de adicionar produto liberando o produto pra ser alterado novamente ao clicar em "Voltar"
@has_permission_decorator('cadastrar_produtos')
def add_produto_redirect(request, slug):
    if request.method == "GET":
        alteracao_produto = Produto.objects.exclude(alterando_produto="0") #Pegando apenas os produtos que tem alguém editando
        if alteracao_produto:
            for alterando in alteracao_produto: #passando por todos os produtos encontrados
                produto_altera_ultimo_acesso = alterando.ultimo_acesso #pegando o último acesso
                if produto_altera_ultimo_acesso != "0":#Caso seja válido entra aqui e pegue o produto
                    alterando.alterando_produto = "0"
                    alterando.ultimo_acesso = "0"
                    alterando.save()
                    return redirect(reverse('add_produto', kwargs={"slug":slug})) 
                    

#Função para a tela de adicionar produto
@has_permission_decorator('cadastrar_produtos')
def add_produto(request, slug):
    opcao = "slug"
    resultado = Consultar_Uma_Comunidade(slug, opcao)
    if resultado[0] != 0:
        id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
        if id_comunidade_comparar_usuario == id_comunidade_usuario:
            if request.method == "GET":
                nome = request.GET.get('nome_produto')
                preco_min = request.GET.get('preco_min')
                preco_max = request.GET.get('preco_max')

                produtos = Produto.objects.filter(nome_comunidade_id=resultado[0])

                Validacao_Alterando_Produto()

                validacao_produtos_filtrados, produtos = Validacao_Produtos_Filtrados(request, slug, produtos, nome, preco_min, preco_max)

                if validacao_produtos_filtrados:
                    return validacao_produtos_filtrados

                nome_produtos = NomeProduto.objects.filter(nome_comunidade_id=resultado[0])
                url_atual = Capturar_Url_Atual_Sem_O_Final(request)

                context = {
                    'nome_produtos': nome_produtos,
                    'produtos': produtos,
                    'slug': slug,
                    'url_atual': url_atual,
                }

                return render(request, 'add_produto.html', context)
            elif request.method == "POST":
                nome = request.POST.get('nome_produto')
                quantidade = request.POST.get('quantidade')
                preco_compra = request.POST.get('preco_compra')
                peso = request.POST.get('peso')
                tipo_peso = request.POST.get('tipo_peso')
                preco_compra = preco_compra.replace(',', '.') # Substitui a vírgula pelo ponto
                preco_venda = 1
                if peso != 0 and peso != "" and peso != None:
                    peso_float = float(peso)
                    if peso_float > 1:
                        messages.add_message(request, messages.ERROR, f'Produto {nome} não pode ter mais de 1.0KG/ML')
                        return redirect(reverse('add_produto', kwargs={"slug":slug}))
                    preco_venda *= peso_float
                nome_produto_original = nome
                acao = "adicionar"
                tabela = "produto"


                with transaction.atomic():
                    resultado = Consultar_Uma_Comunidade(slug, opcao)
                    if resultado[0] != 0:

                        label = nome
                        slugp = slugify(nome + "-" + slug)

                        validacao = Validacoes_Post_Cadastro_Estoque(request, slug, nome, preco_compra, preco_venda, quantidade, slugp, peso)    
                        if validacao:
                            transaction.set_rollback(True)
                            return validacao

                        nome = Capturar_Id_Do_Nome_Do_Produto(nome)

                        num_sequencial = Gerando_Numero_Sequencial(tabela)

                        Cadastro_Estoque(request, nome, label, quantidade, preco_compra, preco_venda, slugp, resultado[0], num_sequencial, peso, tipo_peso)

                        Cadastro_Planilhas_Estoque_E_Atualizacoes_De_Valores(request, slugp, quantidade, preco_compra, preco_venda, acao, resultado[1], peso, 1)

                        messages.add_message(request, messages.SUCCESS, f'Produto {nome_produto_original} Cadastrado com sucesso')
                        return redirect(reverse('add_produto', kwargs={"slug":slug}))
                    else:
                        messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                        return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
            return redirect(reverse('home'))
    else:
        messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
        return redirect(reverse('home'))

#Função para a tela de adicionar Novo nome de Produtos
@has_permission_decorator('cadastrar_produtos')
def add_novonome_produto(request, slug):
    opcao = "slug"
    resultado = Consultar_Uma_Comunidade(slug, opcao)

    if resultado[0] != 0:
        id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
        if id_comunidade_comparar_usuario == id_comunidade_usuario:
            if request.method == "GET":
                    produtos = NomeProduto.objects.filter(nome_comunidade_id=resultado[0])
                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)

                    context = {
                        'produtos': produtos,
                        'slug': slug,
                        'url_atual': url_atual,
                    }

                    return render(request, 'add_novonome_produto.html', context)
            elif request.method == "POST":
                nome_produto = request.POST.get('nome_produto')
                with transaction.atomic():
                    resultado = Consultar_Uma_Comunidade(slug, opcao)
                    produtos = NomeProduto.objects.filter(nome_comunidade_id=resultado[0])

                    validacao_campos = Validacoes_Post_Cadastro_Produtos_Campos_Preenchidos(request, slug, nome_produto)
                    if validacao_campos:
                        transaction.set_rollback(True)
                        return validacao_campos

                    validacao_produto = Cadastrar_Nome_Produto(request, slug, nome_produto, produtos, resultado[0], resultado[4], resultado[5])
                    if validacao_produto:
                        transaction.set_rollback(True)
                        return validacao_produto

                    messages.add_message(request, messages.SUCCESS, f'Produto {nome_produto} cadastrado com sucesso')
                    return redirect(reverse('add_novonome_produto', kwargs={"slug":slug}))
        else:
            messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
            return redirect(reverse('home'))
    else:
        messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
        return redirect(reverse('home'))


#Função para a tela de excluir nome dos Produtos
@has_permission_decorator('excluir_produtos')
def excluir_novonome_produto(request, slug):
    try :#Tente Excluir
        opcao = "id"
        produto = get_object_or_404(NomeProduto, slug=slug)

        id_comunidade_funcionario = produto.nome_comunidade_id #Pegando o ID da comunidade do funcionario
        resultado = Consultar_Uma_Comunidade(id_comunidade_funcionario, opcao)
        if resultado[0] != 0:
            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:
                if resultado[1]:
                    if hasattr(produto, '_excluido'):#Verifica se já foi excluído para não ocorrer repetição de registro no Banco.
                        # se a flag _excluido já está setada, não chama o sinal
                        pass
                    else:#Caso não tenha sido excluído ele chama o registro.
                        novonome_produto_deleted(instance=produto, user=request.user)
                    produto.delete()
                    messages.add_message(request, messages.SUCCESS, 'Produto excluído com sucesso')
                    return redirect(reverse('add_novonome_produto', kwargs={"slug":resultado[1]}))
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))
    except ProtectedError:#Caso não consiga, entre aqui
        messages.add_message(request, messages.ERROR, 'Esse Nome de Produto não pode ser excluído pois possui produtos vinculados')
        return redirect(reverse('add_novonome_produto', kwargs={"slug":resultado[1]}))


#Função para a tela de alterar produto
@has_permission_decorator('editar_produtos')
def produto (request, slug):
    if request.method == "GET":
        opcao = "id"
        produto = Produto.objects.get(slug=slug)
        data = produto.__dict__
        data['nome_produto'] = produto.nome_produto_id
        form = ProdutoForm(initial=data)
        
        id_comunidade_produto = produto.nome_comunidade_id #Pegando o ID da comunidade do produto
        resultado = Consultar_Uma_Comunidade(id_comunidade_produto, opcao)

        if resultado[0] != 0:
            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:
                alterando_produto = produto.alterando_produto
                label = produto.label
                usuario = str(request.user) #Pegando o nome do usuario em formato de string.
                if alterando_produto == "0":
                    ultimo_acesso = Gerar_Token_Com_Tempo_Minutos(5)

                    produto.alterando_produto = usuario #adicionando o nome do usuario à coluna
                    produto.ultimo_acesso = ultimo_acesso #adicionando o contador à coluna
                    produto.save()
                elif alterando_produto == usuario:
                    pass
                else:
                    messages.add_message(request, messages.ERROR, f'O Produto {label} está sendo editado pelo usuário {alterando_produto}')
                    return redirect(reverse('add_produto', kwargs={"slug":resultado[1]}))

                nome_produto = ""

                nomes = NomeProduto.objects.all()#Pegando todos os nomes de produtoso
                if nomes:
                    nomes = nomes.filter(id__contains=produto.nome_produto_id)#Verificando se existem nomes de produto com o nome escolhido
                    nomes = NomeProduto.objects.get(id=produto.nome_produto_id)
                    nome_produto = nomes.nome_produto
                    
                url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                context = {
                    'form': form,
                    'slug': resultado[1],
                    'nome_produto': nome_produto,
                    'url_atual': url_atual,
                }

                return render(request, 'produto.html', context)
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))

    elif request.method == "POST":
        abastecer_quantidade = request.POST.get('abastecer_quantidade')
        preco_compra = request.POST.get('preco_compra')
        tipo_peso = request.POST.get('tipo_peso')

        opcao = "id"

        if abastecer_quantidade == None or not abastecer_quantidade:
            abastecer_quantidade = 0
        with transaction.atomic():
            nome_produto_antigo = Produto.objects.get(slug=slug)
            id_produto_antigo = nome_produto_antigo.id

            id_comunidade_produto = nome_produto_antigo.nome_comunidade_id #Pegando o ID da comunidade do produto
            resultado = Consultar_Uma_Comunidade(id_comunidade_produto, opcao)

            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:

                nome_produto_antigo1 = nome_produto_antigo.__dict__
                nome_produto_antigo1['nome_produto'] = nome_produto_antigo.nome_produto
                nome_produto_antigo1 = nome_produto_antigo1['nome_produto']

                produto = get_object_or_404(Produto, slug=slug) #pegando o slug do produto
                quantidade_atual = produto.quantidade
                tipo_peso_atual = produto.tipo_peso

                alterado_por_ = auth.get_user(request)
                alterado_por = alterado_por_.username

                data_alteracao = Capturar_Ano_E_Hora_Atual()

                preco_compra = preco_compra.replace(',', '.') # Substitui a vírgula pelo ponto
                preco_compra = float(preco_compra)#transformando em float

                if abastecer_quantidade or preco_compra or tipo_peso:
                    if preco_compra == 0 or not preco_compra:
                        messages.add_message(request, messages.ERROR, 'Preço de Compra não pode ser vazio')
                        return redirect(reverse('produto', kwargs={"slug":slug}))

                    if tipo_peso != "KG" and tipo_peso != "ML":
                        messages.add_message(request, messages.ERROR, 'Tipo Peso só pode ser "KG" ou "ML"')
                        return redirect(reverse('produto', kwargs={"slug":slug}))
                    
                    validacao = Registrar_Log_Alteracao_Produto_E_Alterar_Produto(request, resultado[1], slug, id_produto_antigo,abastecer_quantidade, preco_compra, tipo_peso, nome_produto_antigo1, data_alteracao, alterado_por)
                    
                    if validacao:
                        transaction.set_rollback(True)
                        return validacao

                    messages.add_message(request, messages.SUCCESS, (f'Produto {nome_produto_antigo1} atualizado com sucesso'))
                    return redirect(reverse('add_produto', kwargs={"slug":resultado[1]}))
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))


#Função para a tela de excluir produto
@has_permission_decorator('excluir_produtos')
def excluir_produto(request, slug):
    opcao = "id"
    acao = "excluir"
    produto = get_object_or_404(Produto, slug=slug) #pegando o slug do produto

    id_comunidade_produto = produto.nome_comunidade_id #Pegando o ID da comunidade do produto
    id_produto = produto.id
    peso = produto.peso
    nome_produto_p_excel = str(produto.nome_produto)

    resultado = Consultar_Uma_Comunidade(id_comunidade_produto, opcao)
    if resultado[0] != 0:
        id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
        if id_comunidade_comparar_usuario == id_comunidade_usuario:
            opcao = "produto_id"
            with transaction.atomic():
                if hasattr(produto, '_excluido'):#Verifica se já foi excluído para não ocorrer repetição de registro no Banco.
                    # se a flag _excluido já está setada, não chama o sinal
                    pass
                else:#Caso não tenha sido excluído ele chama o registro.
                    produto_deleted(instance=produto, user=request.user)
                produto_ja_foi_vendido = 0

                vendas = Consultar_Uma_Venda(id_produto, opcao)

                if vendas[1] == id_produto:
                    produto_ja_foi_vendido = 1

                #se não tiver venda
                if produto_ja_foi_vendido == 0:
                    Cadastro_Planilhas_Estoque_E_Atualizacoes_De_Valores(request, slug, "sem acao", "sem acao", "sem acao", acao, resultado[1], peso, "sem acao")

                    messages.add_message(request, messages.SUCCESS, 'Produto excluído com sucesso')
                    return redirect(reverse('add_produto', kwargs={"slug":resultado[1]}))
                else:  #Se tiver venda entra aqui
                    if vendas[0] != 0: #caso tenha venda
                        existe_saida_venda = P_Excel.objects.get(nome_produto=nome_produto_p_excel, acao="Saída")
                        if existe_saida_venda.quantidade == 0:
                            Cadastro_Planilhas_Estoque_E_Atualizacoes_De_Valores(request, slug, "sem acao", "sem acao", "sem acao", acao, resultado[1], peso, "sem acao")

                            messages.add_message(request, messages.SUCCESS, 'Produto excluído com sucesso')
                            return redirect(reverse('add_produto', kwargs={"slug":resultado[1]}))
                        transaction.set_rollback(True)
                        messages.add_message(request, messages.ERROR, 'Já existem vendas vinculadas à este produto, contate a administração para a exclusão')
                        return redirect(reverse('add_produto', kwargs={"slug":resultado[1]}))
        else:
            messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
            return redirect(reverse('home'))
    else:
        messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
        return redirect(reverse('home'))


#Função para a tela de cadastrar comunidade
@has_permission_decorator('cadastrar_comunidade')
def cadastrar_comunidade (request):
    if request.method == "GET":
        url_atual = request.path
        context = {
            'url_atual': url_atual
        }
        return render(request, 'cadastrar_comunidade.html', context)
    elif request.method == "POST":
        cnpj = request.POST.get('cnpj')
        tipo = request.POST.get('tipo')
        nome = request.POST.get('nome_comunidade')
        cidade = request.POST.get('cidade')
        responsavel_01 = request.POST.get('responsavel_01')
        celular_01 = request.POST.get('celular_01')
        responsavel_02 = request.POST.get('responsavel_02')
        celular_02 = request.POST.get('celular_02')
        
        with transaction.atomic():
            validacao_comunidades = Validacoes_Cadastro_Comunidades(request, cnpj, tipo, nome, cidade, responsavel_01, celular_01, responsavel_02, celular_02)
            if validacao_comunidades:
                transaction.set_rollback(True)
                return validacao_comunidades
                
            opcao = "nome"
            nome_e_cidade_comunidade = [nome, cidade]
            resultado = Consultar_Uma_Comunidade(nome_e_cidade_comunidade, opcao)
            
            if resultado[0] != 0:
                messages.add_message(request, messages.ERROR, 'Essa comunidade já foi cadastrada nessa mesma cidade, caso isso seja um erro contate a administração')
                return redirect(reverse('cadastrar_comunidade'))
            
            opcao = "cnpj"
            resultado = Consultar_Uma_Comunidade(cnpj, opcao)

            if resultado[0] != 0:
                messages.add_message(request, messages.ERROR, f'Esse CNPJ já foi cadastrado na {resultado[3]}: {resultado[4]} da cidade: {resultado[5]}, caso isso seja um erro contate a administração')
                return redirect(reverse('cadastrar_comunidade'))

            criado_por = auth.get_user(request)

            Cadastrar_Comunidade(cnpj, tipo, nome, cidade, responsavel_01, celular_01, responsavel_02, celular_02, criado_por)

            messages.add_message(request, messages.SUCCESS, (f'Comunidade: {nome} da cidade: {cidade} cadastrada com sucesso'))
            return redirect(reverse('home'))


#Função para a tela de antes de vender produto
@has_permission_decorator('realizar_venda')
def pre_vendas(request, slug):
    opcao = "slug"
    resultado = Consultar_Uma_Comunidade(slug, opcao)
    if resultado[0] != 0:
        id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
        if id_comunidade_comparar_usuario == id_comunidade_usuario:
            if request.method == "GET":
                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                    context = {
                        'slug': slug,
                        'nome_e_cidade_comunidade': slug,
                        'url_atual': url_atual,
                    }
                    return render(request, 'pre_vendas.html', context)
            elif request.method == "POST":
                cpf = request.POST.get('cpf')
                tela = "pre_vendas"

                with transaction.atomic():
                    validacao = validar_cpf(request, cpf, slug, tela)
                    if validacao:
                        transaction.set_rollback(True)
                        return validacao
                    
                    opcao = "cpf"
                    valor = [resultado[0], cpf]

                    resultado_familia = Consultar_Familia(valor, opcao)
                    ativo = resultado_familia[6]
                    ultima_compra = resultado_familia[4]

                    if resultado_familia[0] != 0:
                        if ativo != 0:
                            if ativo == "sim":
                                if ultima_compra is not None and ultima_compra != 0:
                                    
                                    # Data de hoje
                                    hoje = datetime.today().date()

                                    # Verificar se já passaram mais de 7 dias
                                    diferenca = hoje - ultima_compra
                                    if diferenca.days >= 7:
                                        token_venda = Gerar_Token()
                                        Salvando_Novo_Token_Venda_Familia(cpf, token_venda)
                                        
                                        messages.add_message(request, messages.SUCCESS, (f'Esse CPF está autorizado a realizar compra'))
                                        return redirect(reverse('vendas', kwargs={"slug":token_venda}))
                                    else:
                                        messages.add_message(request, messages.ERROR, (f'Esse CPF realizou compra nos últimos 7 dias'))
                                        return redirect(reverse('pre_vendas', kwargs={"slug":slug}))
                                else:
                                    token_venda = Gerar_Token()
                                    Salvando_Novo_Token_Venda_Familia(cpf, token_venda)

                                    messages.add_message(request, messages.SUCCESS, (f'Esse CPF está autorizado a realizar compra'))
                                    return redirect(reverse('vendas', kwargs={"slug":token_venda}))
                            elif ativo == "nao":
                                messages.add_message(request, messages.ERROR, (f'Esse CPF não está ativo na comunidade, caso isso seja um problema, favor contatar a administração'))
                                return redirect(reverse('pre_vendas', kwargs={"slug":slug}))
                        else:
                            messages.add_message(request, messages.ERROR, (f'CPF não encontrado nessa comunidade'))
                            return redirect(reverse('pre_vendas', kwargs={"slug":slug}))
                    else:
                        messages.add_message(request, messages.ERROR, (f'CPF não encontrado nessa comunidade'))
                        return redirect(reverse('pre_vendas', kwargs={"slug":slug}))
        else:
            messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
            return redirect(reverse('home'))
    else:
        messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
        return redirect(reverse('home'))


#Função para a tela de vender produto
@has_permission_decorator('realizar_venda')
def vendas(request, slug):
    if request.method == "GET":
        opcao = "token_venda"
        resultado_familia = Consultar_Familia(slug, opcao)
        if resultado_familia[0] != 0:
            id_familia = resultado_familia[0]
            slug_token_venda_familia = resultado_familia[9]
            nome_cliente_familia = resultado_familia[3]

            opcao = "id"
            resultado = Consultar_Uma_Comunidade(resultado_familia[1], opcao)
            if resultado[0] != 0:
                id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
                if id_comunidade_comparar_usuario == id_comunidade_usuario:
                    id_comunidade = resultado[0]
                    slug_comunidade = resultado[1]

                    BuscaVendas = Q(
                            Q(venda_finalizada=0) & Q(nome_comunidade_id=id_comunidade)     
                    )

                    vendas = Vendas.objects.filter(BuscaVendas)

                    produtos = Produto.objects.filter(nome_comunidade_id=id_comunidade)

                    nome_produtos = NomeProduto.objects.all()
                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                    context = {
                        'nome_produtos':nome_produtos, 
                        'produtos':produtos, 
                        'vendas': vendas, 
                        'slug': slug_comunidade,
                        'slug_token_venda_familia': slug_token_venda_familia,
                        'nome_cliente_familia': nome_cliente_familia,
                        'url_atual': url_atual,
                    }

                    return render(request, 'vendas.html', context)
                else:
                    messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                    return redirect(reverse('home'))
            else:
                messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))

    elif request.method == "POST":
        forma_venda = request.POST.get('forma_venda')
        produtos_selecionados_json = request.POST.get('produtos_selecionados')  # Produtos selecionados

        opcao = "token_venda"
        resultado_familia = Consultar_Familia(slug, opcao)
        nome_cliente = resultado_familia[3]
        cpf = resultado_familia[2]

        opcao = "id"
        resultado = Consultar_Uma_Comunidade(resultado_familia[1], opcao)
        if resultado[0] != 0:
            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:
                id_comunidade = resultado[0]
                slug_comunidade = resultado[1]
                slug_voltar_tela = slug_comunidade

                # Converte a string JSON de produtos selecionados para um objeto Python
                produtos_selecionados = json.loads(produtos_selecionados_json)
                contador_vendas = len(produtos_selecionados)
                tabela = "venda"
                nome_dos_produtos = ""

                with transaction.atomic():
                    num_sequencial = Gerando_Numero_Sequencial(tabela)

                    # Se passar pelas validações, crie o objeto VendasControle contendo o ID venda e salve no banco
                    vendacontrole = Criando_Vendas_Controle(request, nome_cliente, num_sequencial, id_comunidade, contador_vendas, forma_venda)

                    preco_original_venda = 0
                    cont_num_aleatorio = 0
                    preco_total_controle = 0
                    peso_total_controle = 0

                    for produto in produtos_selecionados:
                        nome_produto, preco, peso, peso_item_total, quantidade = Capturar_Valores_Post_Tela_Vendas(produto)

                        try:
                            if not nome_cliente:
                                transaction.set_rollback(True)
                                messages.add_message(request, messages.ERROR, 'Você não preencheu o nome do cliente')
                                return redirect(reverse('vendas', kwargs={"slug":slug})) 

                            # Inicio Declaração de Variáveis da Venda 
                            nome = nome_cliente
                            num_aleatorio = "v" + str(num_sequencial)
                            cont_num_aleatorio += 1
                            label = nome_produto
                            label_vendas_get = nome_produto

                            nome_cliente = nome_cliente
                            nome_cliente = unidecode.unidecode(f'{nome_cliente}')
                            nome_cliente = str(nome_cliente)
                            
                            quantidade_estoque_produto = 0
                            id_produto = 0
                            id_nome_produto = 0
                            produto = 0
                            preco_compra_estoque_produto = 0
                            preco_venda_estoque_produto = 0

                            data_atual = Capturar_Ano_E_Hora_Atual()
                            
                            label = slugify(nome_produto + "-" + data_atual)
                            slugp = slugify(nome_produto + "-" + data_atual + "-" + num_aleatorio)

                            valor = [id_comunidade, nome_produto]
                            # Fim Declaração de Variáveis da Venda 

                            labelf = Vendas.objects.filter(label_vendas=label)#Verificando se está sendo preenchido produtos duplicados na mesma venda.

                            if labelf:
                                transaction.set_rollback(True) #Desfazer alterações caso haja erro em alguma.
                                messages.add_message(request, messages.ERROR, 'Confira os produtos, parecem duplicados')
                                return redirect(reverse('vendas', kwargs={"slug":slug}))

                            resultado_nome_produto = Consultar_Nome_Dos_Produtos(valor, opcao)
                            id_nome_produto = resultado_nome_produto[0]

                            if id_nome_produto != 0:
                                valor = [id_nome_produto, nome_produto, id_comunidade]
                                resultado_produto = Consultar_Dados_Dos_Produtos(valor, opcao)
                                id_produto = resultado_produto[0]
                                slug_produto = resultado_produto[1]

                                if id_produto != 0:
                                    quantidade_estoque_produto = resultado_produto[4]
                                    preco_compra_estoque_produto = resultado_produto[5]
                                    preco_venda_estoque_produto = preco
                                    nome_produto_str_ = resultado_produto[3]
                                    nome_produto_str = str(nome_produto_str_)

                                    if quantidade_estoque_produto <= 0:
                                        transaction.set_rollback(True)
                                        messages.add_message(request, messages.ERROR, 'Não temos esse Produto no estoque no momento')
                                        return redirect(reverse('vendas', kwargs={"slug":slug}))              
                                else:
                                    transaction.set_rollback(True)
                                    messages.add_message(request, messages.ERROR, 'Algo de errado aconteceu com sua venda, caso persista, contate a administração.')
                                    return redirect(reverse('vendas', kwargs={"slug":slug}))  
                            else:
                                transaction.set_rollback(True)
                                messages.add_message(request, messages.ERROR, 'Algo de errado aconteceu com sua venda, caso persista, contate a administração.')
                                return redirect(reverse('vendas', kwargs={"slug":slug}))                  
                        except Exception as e:
                            transaction.set_rollback(True)
                            messages.add_message(request, messages.ERROR, 'Ocorreu um erro ao processar a venda, fale com a administração')
                            return redirect(reverse('vendas', kwargs={"slug":slug}))

                        validacao_forma_venda_e_qtd, preco_venda_estoque_produto, preco_venda_estoque_produto_real, preco_venda_total = Valida_Forma_Venda_E_Quantidade(request, slug, forma_venda, quantidade, quantidade_estoque_produto, nome_produto, id_produto)

                        if validacao_forma_venda_e_qtd:
                            transaction.set_rollback(True)
                            return validacao_forma_venda_e_qtd

                        # Inicio Declaração Variaveis
                        preco_total_controle += float(preco_venda_total)
                        peso_total_controle += float(peso_item_total)
                        preco_venda = preco_venda_estoque_produto_real
                        preco_original_venda += float(preco_venda_total)
                        # Fim Declaração Variaveis

                        # Se passar pelas validações, crie o objeto Venda e salve no banco
                        Criando_Vendas(request, id_nome_produto, quantidade, peso, peso_item_total, forma_venda, preco_compra_estoque_produto, preco_venda, preco_venda_total, slugp, id_comunidade, label, label_vendas_get, id_produto, nome_cliente, vendacontrole)

                        id_user, id_comunidade_usuario = Capturar_Id_E_Comunidade_Do_Usuario(request)

                        acao = "vender"
                        Cadastro_Planilhas_Estoque_E_Atualizacoes_De_Valores(request, slug_produto, quantidade, preco_compra_estoque_produto, preco_venda_estoque_produto, acao, slug_comunidade, peso, id_produto)

                        nome_dos_produtos = Capturar_Nome_Dos_Produtos(nome_dos_produtos, nome_produto)

                    # Se passar pelas validações, pega o objeto VendasControle contendo o ID venda gerado lá em cima#
                    Atualiza_Venda_Controle(num_sequencial, preco_total_controle, nome_dos_produtos, peso_total_controle)

                    Salvando_Novo_Token_Venda_Familia(cpf, "reset")

                    messages.add_message(request, messages.SUCCESS, 'Venda Cadastrada com sucesso')
                    return redirect(reverse('pre_vendas', kwargs={"slug":slug_voltar_tela}))
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))


#Função para a tela de consultar vendas geral
@has_permission_decorator('consultar_venda')
def consultar_vendas_geral(request, slug):
    if request.method == "GET":
        nome_cliente = request.GET.get('nome_cliente_filtro')
        nome = request.GET.get('nome_produto_filtro')
        funcionario = request.GET.get('funcionario')
        get_dt_start = request.GET.get('dt_start')
        get_dt_end = request.GET.get('dt_end')

        opcao = "slug"
        resultado = Consultar_Uma_Comunidade(slug, opcao)
        if resultado[0] != 0:
            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:
                    id_comunidade = resultado[0]

                    BuscaVendas = Q(
                            Q(venda_finalizada=0) & Q(nome_comunidade_id=id_comunidade)     
                    )

                    vendas = VendasControle.objects.filter(BuscaVendas)
                    validacao, page = Get_Paginacao_Vendas_Controle(request, slug, nome_cliente, nome, get_dt_start, get_dt_end, funcionario, vendas)
                    if validacao:
                        return validacao
                        
                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                    context = {
                        'vendas': vendas, 
                        'page': page,
                        'slug': slug,
                        'url_atual': url_atual,
                    }

                    return render(request, 'consultar_vendas_geral.html', context)
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))


#Função para a tela de consultar venda
@has_permission_decorator('consultar_venda')
def consultar_vendas(request, slug):
    if request.method == "GET":
        opcao = "id_venda_consultar_vendas"
        vendas_pra_pegar_id = Consultar_Uma_Venda(slug, opcao)
        if vendas_pra_pegar_id:
            id_comunidade = vendas_pra_pegar_id[13]

            BuscaVendas = Q(
                Q(venda_finalizada=0) & Q(id_venda_id=slug) & Q(nome_comunidade_id=id_comunidade)     
            ) 
            vendas = Vendas.objects.filter(BuscaVendas)

            opcao = "id"
            resultado = Consultar_Uma_Comunidade(id_comunidade, opcao)

            if resultado[0] != 0:
                id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
                if id_comunidade_comparar_usuario == id_comunidade_usuario:
                    id_comunidade = resultado[0]
                    slug_comunidade = resultado[1]

                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                    context = {
                        'vendas': vendas,
                        'slug': slug_comunidade,
                        'url_atual': url_atual,
                    }

                    return render(request, 'consultar_vendas.html', context)
                else:
                    messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                    return redirect(reverse('home'))
            else:
                messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home')) 


#Função para a tela de consultar venda
@has_permission_decorator('consultar_venda')
def consultar_vendas_finalizadas(request, slug):
    if request.method == "GET":
        opcao = "id_venda_consultar_vendas_finalizadas"
        vendas_pra_pegar_id = Consultar_Uma_Venda(slug, opcao)
        if vendas_pra_pegar_id:
            id_comunidade = vendas_pra_pegar_id[13]

            BuscaVendas = Q(
                Q(venda_finalizada=1) & Q(id_venda_id=slug) & Q(nome_comunidade_id=id_comunidade)     
            ) 
            vendas = Vendas.objects.filter(BuscaVendas)

            opcao = "id"
            resultado = Consultar_Uma_Comunidade(id_comunidade, opcao)

            if resultado[0] != 0:
                id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
                if id_comunidade_comparar_usuario == id_comunidade_usuario:
                    id_comunidade = resultado[0]
                    slug_comunidade = resultado[1]

                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                    context = {
                        'vendas': vendas,
                        'slug': slug_comunidade,
                        'url_atual': url_atual,
                    }

                    return render(request, 'consultar_vendas_finalizadas.html', context)
                else:
                    messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                    return redirect(reverse('home'))
            else:
                messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home')) 


#Função para a tela de vender produto
@has_permission_decorator('finalizar_venda', 'cancelar_venda')
def vendas_finalizadas(request, slug):
    if request.method == "GET":
        nome_cliente = request.GET.get('nome_cliente_filtro')
        nome = request.GET.get('nome_produto_filtro')
        funcionario = request.GET.get('funcionario')
        get_dt_start = request.GET.get('dt_start')
        get_dt_end = request.GET.get('dt_end')

        opcao = "slug"
        resultado = Consultar_Uma_Comunidade(slug, opcao)
        if resultado[0] != 0:
            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:
                id_comunidade = resultado[0]

                BuscaVendasFinalizadas = Q(
                        Q(venda_finalizada=1) & Q(nome_comunidade_id=id_comunidade)     
                )

                vendas = VendasControle.objects.filter(BuscaVendasFinalizadas)

                validacao, page = Get_Paginacao_Vendas_Finalizadas(request, slug, nome_cliente, nome, get_dt_start, get_dt_end, funcionario, vendas)
                if validacao:
                    return validacao

                produtos = Produto.objects.filter(nome_comunidade_id=id_comunidade)

                nome_produtos = NomeProduto.objects.all()

                #o sinal de "~" antes significa negação
                BuscaVendasNaoFinalizadas = Q(
                        ~Q(venda_finalizada=1) & Q(nome_comunidade_id=id_comunidade)     
                )

                all_vendas = VendasControle.objects.filter(BuscaVendasNaoFinalizadas).order_by('-id_venda') #Vendas mais novas primeiro, ordenadas pelo ID
                url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                context = {
                    'nome_produtos':nome_produtos, 
                    'produtos':produtos, 
                    'vendas': vendas, 
                    'page': page,
                    'slug': slug,
                    'id_comunidade': id_comunidade,
                    'all_vendas': all_vendas,
                    'url_atual': url_atual,
                }

                return render(request, 'vendas_finalizadas.html', context)
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home')) 


#Função para a tela de visualizar vendas
@has_permission_decorator('consultar_venda')
def visualizar_vendas (request, slug):
    if request.method == "GET":
        venda = Vendas.objects.filter(slug=slug)
        if venda:
            venda = Vendas.objects.get(slug=slug)
            venda_finalizada = venda.venda_finalizada
            slug_venda = venda.id_venda_id
            if venda_finalizada == 0:
                # Inicio Declaração de Variaveis
                id_comunidade = venda.nome_comunidade_id
                opcao = "id"
                resultado = Consultar_Uma_Comunidade(id_comunidade, opcao)
                if resultado[0] != 0:
                    id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
                    if id_comunidade_comparar_usuario == id_comunidade_usuario:
                        slug_comunidade = resultado[1]

                        id_nome_produto = venda.nome_produto_id

                        BuscaProduto = Q( #Fazendo o Filtro com Busca Q para a tabela Vendas
                            Q(nome_comunidade_id=id_comunidade) & Q(nome_produto_id=id_nome_produto) 
                        )

                        produto = Produto.objects.get(BuscaProduto)
                        cod_produto = produto.cod_produto

                        data_criacao = venda.data_criacao
                        quantidade_venda = venda.quantidade
                        vendido_por = venda.criado_por
                        forma_venda = venda.forma_venda
                        nome_cliente = venda.nome_cliente

                        preco_compra = venda.preco_compra#Pegando preco_compra

                        preco_venda = venda.preco_venda#Pegando preco_venda
                        preco_venda_total = venda.preco_venda_total#Pegando preco_venda total
                        nome_produto = ""
                        # Fim Declaração de Variaveis

                        if preco_compra or preco_venda:
                            preco_compra = "R$ " + str(preco_compra)
                            preco_venda = "R$ " + str(preco_venda)
                            preco_venda_total = "R$ " + str(preco_venda_total)

                        nomes = NomeProduto.objects.all()#Pegando todos os nomes de produtos
                        if nomes:
                            nomes = nomes.filter(id__contains=venda.nome_produto_id)#Verificando se existem nomes de produto com o nome escolhido
                            nomes = NomeProduto.objects.get(id=venda.nome_produto_id)
                            nome_produto = nomes.nome_produto

                        url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                        context = {
                            'slug_venda': slug_venda,
                            'slug': slug_comunidade,
                            'vendido_por':vendido_por,
                            'preco_venda':preco_venda,
                            'preco_compra':preco_compra,
                            'quantidade_venda':quantidade_venda,
                            'data_criacao':data_criacao,
                            'nome_produto':nome_produto,
                            'forma_venda':forma_venda,
                            'nome_cliente':nome_cliente,
                            'venda_finalizada':venda_finalizada,
                            'preco_venda_total':preco_venda_total,
                            'cod_produto': cod_produto,
                            'url_atual': url_atual,
                        }
                        return render(request, 'visualizar_vendas.html', context)
                    else:
                        messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                        return redirect(reverse('home'))
                else:
                    messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                    return redirect(reverse('home'))
            else:
                messages.add_message(request, messages.ERROR, 'Essa venda já foi finalizada')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))


#Função para a tela de visualizar vendas
@has_permission_decorator('consultar_venda')
def visualizar_vendas_finalizadas (request, slug):
    if request.method == "GET":
        venda = Vendas.objects.filter(slug=slug)
        if venda:
            venda = Vendas.objects.get(slug=slug)
            venda_finalizada = venda.venda_finalizada
            slug_venda = venda.id_venda_id
            if venda_finalizada == 1:
                # Inicio Declaração de Variaveis
                id_comunidade = venda.nome_comunidade_id
                opcao = "id"
                resultado = Consultar_Uma_Comunidade(id_comunidade, opcao)
                if resultado[0] != 0:
                    id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
                    if id_comunidade_comparar_usuario == id_comunidade_usuario:
                        slug_comunidade = resultado[1]

                        id_nome_produto = venda.nome_produto_id

                        BuscaProduto = Q( #Fazendo o Filtro com Busca Q para a tabela Vendas
                            Q(nome_comunidade_id=id_comunidade) & Q(nome_produto_id=id_nome_produto) 
                        )

                        produto = Produto.objects.get(BuscaProduto)
                        cod_produto = produto.cod_produto

                        data_criacao = venda.data_criacao
                        quantidade_venda = venda.quantidade
                        vendido_por = venda.criado_por
                        forma_venda = venda.forma_venda
                        nome_cliente = venda.nome_cliente

                        preco_compra = venda.preco_compra#Pegando preco_compra

                        preco_venda = venda.preco_venda#Pegando preco_venda
                        preco_venda_total = venda.preco_venda_total#Pegando preco_venda total
                        nome_produto = ""
                        # Fim Declaração de Variaveis

                        if preco_compra or preco_venda:
                            preco_compra = "R$ " + str(preco_compra)
                            preco_venda = "R$ " + str(preco_venda)
                            preco_venda_total = "R$ " + str(preco_venda_total)

                        nomes = NomeProduto.objects.all()#Pegando todos os nomes de produtos
                        if nomes:
                            nomes = nomes.filter(id__contains=venda.nome_produto_id)#Verificando se existem nomes de produto com o nome escolhido
                            nomes = NomeProduto.objects.get(id=venda.nome_produto_id)
                            nome_produto = nomes.nome_produto

                        url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                        context = {
                            'slug_venda': slug_venda,
                            'slug': slug_comunidade,
                            'vendido_por':vendido_por,
                            'preco_venda':preco_venda,
                            'preco_compra':preco_compra,
                            'quantidade_venda':quantidade_venda,
                            'data_criacao':data_criacao,
                            'nome_produto':nome_produto,
                            'forma_venda':forma_venda,
                            'nome_cliente':nome_cliente,
                            'venda_finalizada':venda_finalizada,
                            'preco_venda_total':preco_venda_total,
                            'cod_produto': cod_produto,
                            'url_atual': url_atual,
                        }
                        return render(request, 'visualizar_vendas_finalizadas.html', context)
                    else:
                        messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                        return redirect(reverse('home'))
                else:
                    messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                    return redirect(reverse('home'))
            else:
                messages.add_message(request, messages.ERROR, 'Essa venda ainda está em andamento')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))


#Função para a tela de conferir os itens da venda antes de finalizar
@has_permission_decorator('finalizar_venda', 'cancelar_venda')
def conferir_vendas_geral (request, slug):
    if request.method == "GET":
        opcao = "id_venda_consultar_vendas"
        vendas_pra_pegar_id = Consultar_Uma_Venda(slug, opcao)
        if vendas_pra_pegar_id:
            id_comunidade = vendas_pra_pegar_id[13]
            forma_venda = vendas_pra_pegar_id[3]

            BuscaVendas = Q(
                Q(venda_finalizada=0) & Q(id_venda_id=slug) & Q(nome_comunidade_id=id_comunidade)     
            ) 
            conferir_vendas = Vendas.objects.filter(BuscaVendas)

            opcao = "id"
            resultado = Consultar_Uma_Comunidade(id_comunidade, opcao)

            if resultado[0] != 0:
                id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
                if id_comunidade_comparar_usuario == id_comunidade_usuario:
                    id_comunidade = resultado[0]
                    slug_comunidade = resultado[1]

                    url_atual = Capturar_Url_Atual_Sem_O_Final(request)
                    context = {
                        'conferir_vendas': conferir_vendas,
                        'slug': slug_comunidade,
                        'url_atual': url_atual,
                        'forma_venda':forma_venda,
                    }

                    return render(request, 'conferir_vendas_geral.html', context)
                else:
                    messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                    return redirect(reverse('home'))
            else:
                messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))

    if request.method == "POST":
        with transaction.atomic():
            precos_venda_total = request.POST.getlist('preco-venda-total')
            label_vendas_get = request.POST.getlist('label_vendas_get')
            quantidades = request.POST.getlist('quantidade')
            forma_pagamento_nova = request.POST.get('alterar_forma_venda')

            contador_qtd_alterada = 0
            contador_produtos = len(precos_venda_total)

            # Processar os valores obtidos
            preco_total_controle = 0
            preco_original_venda = 0
            nome_dos_produtos = ""
            valor_total = 0

            for preco_total in precos_venda_total:
                valor_total += float(preco_total)

            valor_total_str = Verificando_Digito_Final_Preco(valor_total)

            if valor_total > 5:
                messages.add_message(request, messages.ERROR, f'O Preço total dos produtos não pode ultrapassar R$5,00, você tentou alterar para: R${valor_total_str}')
                return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slug}))

            combinacao = itertools.zip_longest(quantidades, precos_venda_total, label_vendas_get, fillvalue=0)
            for quantidade, preco, label in combinacao:
                preco = float(preco.replace(',', '.'))
                quantidade = int(quantidade)

                label_da_venda = label
                preco_venda_str = Verificando_Digito_Final_Preco(preco)

                if preco > 1:
                    messages.add_message(request, messages.ERROR, f'O Preço de cada produto não pode ultrapassar R$1,00, o produto: "{label_da_venda}" ultrapassou esse valor, você tentou alterar para: R${preco_venda_str}')
                    return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slug}))

                Busca = Q(
                    Q(id_venda=slug) & Q(label_vendas_get=label_da_venda)
                )
                id_da_venda_ = Vendas.objects.get(Busca)
                slug_venda_atual = id_da_venda_.slug
                id_da_venda = id_da_venda_.id_venda
                quantidade_antiga_da_venda = id_da_venda_.quantidade
                nome_comunidade_id = id_da_venda_.nome_comunidade_id

                valor = [nome_comunidade_id, label_da_venda]
                opcao = "id"
                resultado_nome_produto = Consultar_Nome_Dos_Produtos(valor, opcao)
                id_nome_produto = resultado_nome_produto[0]

                resultado = Consultar_Uma_Comunidade(nome_comunidade_id, opcao)
                slug_comunidade = resultado[1]

                if id_nome_produto != 0:
                    valor = [id_nome_produto, label_da_venda, nome_comunidade_id]
                    resultado_produto = Consultar_Dados_Dos_Produtos(valor, opcao)

                    id_produto = resultado_produto[0]
                    slug_produto = resultado_produto[1]
                    preco_compra_estoque_produto = resultado_produto[5]
                    preco_venda_estoque_produto = resultado_produto[6]
                    quantidade_estoque_produto = resultado_produto[4]

                    tabela_produto = Produto.objects.get(id=id_produto)

                if quantidade <= 0:
                    transaction.set_rollback(True)
                    messages.add_message(request, messages.ERROR, 'Quantidade não pode ser 0')
                    return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slug}))
                elif ((quantidade_estoque_produto + quantidade_antiga_da_venda) - quantidade) < 0:
                    transaction.set_rollback(True)
                    messages.add_message(request, messages.ERROR, f'A quantidade atual em estoque do produto {label_da_venda} é menor do que o aumento solicitado, quantidade atual ({quantidade_estoque_produto}), aumento solicitado ({quantidade-quantidade_antiga_da_venda})')
                    return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slug}))
                else:
                    if quantidade < quantidade_antiga_da_venda:
                        tabela_produto.quantidade += quantidade_antiga_da_venda-quantidade
                        tabela_produto.save()
                    elif quantidade == quantidade_antiga_da_venda:
                        contador_qtd_alterada += 1
                    else:
                        tabela_produto.quantidade += quantidade_antiga_da_venda
                        tabela_produto.quantidade -= quantidade
                        tabela_produto.save()

                    quantidade_estoque_produto = quantidade_estoque_produto - quantidade
                    produto = Produto.objects.filter(id=id_produto) #Buscando produto pelo ID encontrado lá em cima
                    if produto:
                        preco_compra_estoque_produto = float(preco_compra_estoque_produto)
                        preco = float(preco)
                        preco_venda_total = preco
                        trocas_de_quantidade = quantidade_antiga_da_venda - quantidade
                        if trocas_de_quantidade <= 0:
                            trocas_de_quantidade = quantidade - quantidade_antiga_da_venda


                    preco_total_controle += preco_venda_total

                    preco_original_venda += float(preco_venda_total)
                    preco_original = float(preco_venda_total)

                    # Se passar pelas validações, altere o objeto Venda e salve no banco
                    validacao, nova_quantidade, quantidade_antes_de_trocar, forma_pagamento_nova = Alterar_Quantidade_E_Preco_Da_Venda_E_Salvar_No_Banco(request, slug, id_da_venda_, quantidade, preco_venda_total, preco_original, forma_pagamento_nova, label_da_venda)

                    if validacao:
                        return validacao

            #Se passar pelas validações, pega o objeto VendasControle contendo o ID venda gerado lá em cima
            conferir_alteracoes = Conferir_Alteracoes_E_Venda_Controle(request, slug, id_da_venda, forma_pagamento_nova, contador_qtd_alterada, contador_produtos, preco_total_controle, preco_original_venda)
            if conferir_alteracoes:
                return conferir_alteracoes

            data_alteracao = Capturar_Ano_E_Hora_Atual()

            quantidade = [quantidade_antes_de_trocar, nova_quantidade, data_alteracao]
            acao = "alterar"

            Cadastro_Planilhas_Estoque_E_Atualizacoes_De_Valores(request, "sem acao", quantidade, "sem acao", "sem acao", acao, slug_comunidade, "sem acao", id_produto)
            
            messages.add_message(request, messages.SUCCESS, 'Venda alterada com sucesso')
            return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))


#Função para a tela de excluir venda antes da finalização
@has_permission_decorator('finalizar_venda', 'cancelar_venda')
def excluir_venda(request, slug):
    with transaction.atomic():
        venda = Vendas.objects.get(slug=slug)
        slugconferir_venda = venda.id_venda
        conferir_venda = VendasControle.objects.get(slug=slugconferir_venda)
        quantidade_estornar = venda.quantidade
        produto_venda = venda.label_vendas_get
        nome_comunidade_id = venda.nome_comunidade_id
        preco_total = venda.preco_venda_total

        opcao = "id_venda"
        resultado_venda_controle = Consultar_Venda_Controle(slugconferir_venda, opcao)
        id_comunidade = resultado_venda_controle[2]
        label_vendas_get_controle = resultado_venda_controle[15]

        opcao = "id"
        resultado_comunidade = Consultar_Uma_Comunidade(id_comunidade, opcao)
        slug_da_comunidade = resultado_comunidade[1]

        label_vendas_get = remover_palavra(label_vendas_get_controle, produto_venda)

        if hasattr(venda, '_excluido'):#Verifica se já foi excluído para não ocorrer repetição de registro no Banco.
            # se a flag _excluido já está setada, não chama o sinal
            pass
        else:#Caso não tenha sido excluído ele chama o registro.
            vendas_deleted(sender=Vendas, instance=venda, user=request.user)
        if venda:
            BuscaProduto = Q(
                Q(label=produto_venda) & Q(nome_comunidade_id=id_comunidade)     
            ) 
            produto_cancelar = Produto.objects.filter(BuscaProduto)
            if produto_cancelar:
                conferir_venda.novo_preco_venda_total -= preco_total #Retirando o preço do item removido, da tabela de venda controle
                conferir_venda.valor_cancelado += preco_total #Somando o valor do item cancelado
                conferir_venda.falta_editar -= 1
                conferir_venda.label_vendas_get = label_vendas_get
                conferir_venda.save()

                produto_cancelar = Produto.objects.get(BuscaProduto)
                id_produto = produto_cancelar.id #Pegando ID do produto onde deve estornar a quantidade NÃO VENDIDA
                quantidade_atual_produto = produto_cancelar.quantidade #Pegando a quantidade atual do produto

                produto_cancelar.quantidade = quantidade_atual_produto + quantidade_estornar #devolvendo a quantidade NÃO VENDIDA ao produto
                produto_cancelar.save()

                venda.delete() #deletando a venda não realizada do banco.

                #Resolvendo questões da vendas controle

                Busca = Q( #Fazendo o Filtro com Busca Q
                    Q(slug=slugconferir_venda) & Q(falta_editar=0)     
                )

                venda_controle_pos_delete = VendasControle.objects.filter(Busca)

                if venda_controle_pos_delete:#Se entrar aqui é porque todos os itens já foram editados
                    conferir_venda.venda_finalizada = 1 #Alterando venda_finalizada para 1
                    conferir_venda.alteracoes_finalizadas = True
                    conferir_venda.save()

                    venda_pos_delete = Vendas.objects.filter(id_venda=slug)
                    if not venda_pos_delete: #Deletando caso não exista mais compras e essa tenha sido a última à ser deletada
                        vendas_geral_deleted(sender=VendasControle, instance=conferir_venda, user=request.user)#Enviando a venda deletada para o log em vendas_geral_deleted no signals
                        conferir_venda.delete()

                    messages.add_message(request, messages.SUCCESS, 'Venda Cancelada com sucesso')
                    return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_da_comunidade}))

                if not venda_controle_pos_delete:#Se entrar aqui é porque algum item ainda não foi editado, então a compra não será excluída do controle 
                    messages.add_message(request, messages.SUCCESS, 'Venda Cancelada com sucesso')
                    return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slugconferir_venda}))

            else:
                venda.delete()
                messages.add_message(request, messages.SUCCESS, 'Venda Cancelada com sucesso')
                return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slugconferir_venda}))
        else:
            messages.add_message(request, messages.ERROR, 'Houve um erro no cancelamento dessa venda, caso persista contate o administrador')
            return redirect(reverse('conferir_vendas_geral', kwargs={"slug":slugconferir_venda}))


#Função para a tela de excluir venda geral antes da finalização
@has_permission_decorator('finalizar_venda', 'cancelar_venda')
def excluir_venda_geral(request, slug):
    with transaction.atomic():
        Busca = Q(
            Q(id_venda=slug) & Q(modificado=False)     
        ) 
        vendas = Vendas.objects.filter(Busca) #Pegando todas as vendas com o mesmo id_venda e que ainda não foram modificados
        vendas_geral = Vendas.objects.filter(id_venda=slug) #Pegando todas as vendas com o mesmo id_venda
        vendas_geralzao = vendas_geral.count() #Pegando a quantidade de linhas da Query acima

        opcao = "id_venda"
        resultado_venda_controle = Consultar_Venda_Controle(slug, opcao)
        id_comunidade = resultado_venda_controle[2]

        opcao = "id"
        resultado_comunidade = Consultar_Uma_Comunidade(id_comunidade, opcao)
        slug_da_comunidade = resultado_comunidade[1]
        
        conferir_venda = VendasControle.objects.get(slug=slug) #pegando a compra com o mesmo id_venda
        quantidade_estornar = [] 
        produto_venda = [] 
        nome_comunidade_id = []
        preco_total = []
        valor_total_cancelado = 0
        for venda in vendas:#Loop For para colocar quantidade, nome do produto, nome_comunidade_id e preco dentro de arrays.
            quantidade_estornar.append(venda.quantidade)
            produto_venda.append(venda.label_vendas_get)
            nome_comunidade_id.append(venda.nome_comunidade_id)
            preco_total.append(venda.preco_venda_total)    
            valor_total_cancelado += venda.preco_venda_total
            vendas_deleted(sender=Vendas, instance=venda, user=request.user)#Enviando cada venda deletada para o log em vendas_deleted no signals
            venda.delete() #Deletando cada venda
        if vendas:
            for produto_label, quantidade, preco in zip(produto_venda, quantidade_estornar, preco_total):#Devolvendo o produto ao estoque após deletar.
                BuscaProduto = Q(
                    Q(label=produto_label) & Q(nome_comunidade_id=id_comunidade)     
                ) 
                produto_estornar = Produto.objects.get(BuscaProduto)
                id_produto = produto_estornar.id
                quantidade_a_estornar = produto_estornar.quantidade
                produto_estornar.quantidade =  quantidade_a_estornar + quantidade
                produto_estornar.save()

            if vendas_geralzao == vendas.count():#Se a quantidade de linhas não modificadas for a mesma da consulta apenas pelo id_venda então delete tudo.
                vendas_geral_deleted(sender=VendasControle, instance=conferir_venda, user=request.user)#Enviando a venda deletada para o log em vendas_geral_deleted no signals
                conferir_venda.delete()#deletando o controle da venda cancelada do banco.
            else:
                conferir_venda.venda_finalizada = 1 #Alterando venda_finalizada para 1
                conferir_venda.alteracoes_finalizadas = True
                conferir_venda.valor_cancelado += valor_total_cancelado
                conferir_venda.novo_preco_venda_total -= valor_total_cancelado
                conferir_venda.save()

            messages.add_message(request, messages.SUCCESS, 'Venda Cancelada com sucesso')
            return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_da_comunidade}))
        else:
            messages.add_message(request, messages.ERROR, 'Houve um erro no cancelamento dessa venda, caso persista contate o administrador')
            return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_da_comunidade}))


#Função para a tela de confirmar venda
@has_permission_decorator('finalizar_venda', 'cancelar_venda')
def confirmar_venda_geral(request, slug):
    with transaction.atomic():
        valor_pago = request.GET.get('valor_pago')
        troco = request.GET.get('troco')
        quantidade_parcelas = request.GET.get('quantidade_parcelas')

        vendas_geral = Vendas.objects.filter(id_venda=slug) #Pegando todas as vendas com o mesmo id_venda
        vendas_geralzao = vendas_geral.count() #Pegando a quantidade de linhas da Query acima

        opcao = "id_venda"
        resultado_venda_controle = Consultar_Venda_Controle(slug, opcao)
        valor_a_pagar = resultado_venda_controle[8]
        forma_venda = resultado_venda_controle[13]
        id_comunidade = resultado_venda_controle[2]

        opcao = "id"
        resultado_comunidade = Consultar_Uma_Comunidade(id_comunidade, opcao)

        if resultado_comunidade[0] != 0:
            id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado_comunidade[0])
            if id_comunidade_comparar_usuario == id_comunidade_usuario:
                slug_comunidade = resultado_comunidade[1]
                
                if valor_pago == "0" or valor_pago == "" or valor_pago == 0:
                    if forma_venda == "Dinheiro":
                        messages.add_message(request, messages.ERROR, 'Valor Pago não pode ser vazio')
                        return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))
                    else:
                        valor_pago = valor_a_pagar

                if forma_venda == "Crédito":
                    if quantidade_parcelas.isdigit(): 
                        quantidade_parcelas = int(quantidade_parcelas) 
                        if quantidade_parcelas <= 0:
                                messages.add_message(request, messages.ERROR, 'Quantidade de parcelas não pode ser vazio, igual ou menor que zero para crédito')
                                return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))
                        if quantidade_parcelas > 0:
                            pass
                        else:
                            messages.add_message(request, messages.ERROR, 'Você deve preencher um número para quantidade de parcelas')
                            return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))
                    else:
                        messages.add_message(request, messages.ERROR, 'Quantidade de Parcelas deve conter apenas números')#Verificando se contém apenas números
                        return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade})) 
                else:
                    quantidade_parcelas = 0

                valor_pago = str(valor_pago).replace(',', '.') # Substitui a vírgula pelo ponto
                valor_pago = float(valor_pago)#Transformando em float
                valor_pago = Decimal(valor_pago)#Transformando em Decimal

                troco = str(troco).replace(',', '.') # Substitui a vírgula pelo ponto
                troco = float(troco)#Transformando em float
                troco = Decimal(troco)#Transformando em Decimal

                data_alteracao = Capturar_Ano_E_Hora_Atual()

                if resultado_venda_controle[0] != 0:
                    opcao = "filtro-id-venda"
                    resultado_venda = Consultar_Uma_Venda(slug, opcao)
                    for venda in resultado_venda:#Loop For para colocar quantidade, nome do produto, e preco dentro de arrays.
                        preco_total = venda.preco_venda_total
                        id_venda_atual = venda.id

                        if valor_pago < valor_a_pagar:
                            messages.add_message(request, messages.ERROR, 'O Valor pago não pode ser menor que o valor à pagar')
                            return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))
                        if troco < 0:
                            messages.add_message(request, messages.ERROR, 'Troco não pode ser negativo')
                            return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))

                        opcao = "filtro-id-venda"
                        resultado_venda_controle_filter = Consultar_Venda_Controle(slug, opcao)
                        resultado_venda_controle_filter.novo_preco_venda_total -= preco_total #Retirando o preço do item confirmado, da tabela de vendacontrole
                        resultado_venda_controle_filter.falta_editar -= 1
                        resultado_venda_controle_filter.valor_pago += preco_total #Somando o valor do item confirmado (pago)
                        resultado_venda_controle_filter.save()

                        dia = Capturar_Data_Em_Formato_Data()

                        confirmar_venda = True
                        venda_antiga = None
                        venda_antiga = Vendas.objects.get(id=id_venda_atual)

                        campos_alteracao = []
                        if venda_antiga:
                            venda_antiga.venda_finalizada = str(venda_antiga.venda_finalizada)

                            if venda_antiga.venda_finalizada != confirmar_venda:
                                campos_alteracao.append('venda_finalizada')                  

                            venda.venda_finalizada = confirmar_venda #Confirmando a venda
                            venda.modificado = True
                            venda.alterado_por = request.user.username
                            venda.data_criacao = data_alteracao
                            venda.dia = dia
                            venda.data_alteracao = data_alteracao
                            venda.save()

                            #Atualizando a data da venda para a data que foi finalizada.
                            vendas_ = Vendas.objects.filter(id_venda=slug)
                            for venda in vendas_:
                                venda.dia = dia
                                venda.data_criacao = data_alteracao
                                venda.save()

                            venda_nova = None
                            venda_nova = Vendas.objects.get(id=id_venda_atual)
                            if campos_alteracao:
                                valores_antigos = []
                                valores_novos = []
                                for campo in campos_alteracao:
                                    valor_antigo = getattr(venda_antiga, campo)
                                    valor_novo = getattr(venda_nova, campo)
                                    valores_antigos.append(f'{campo}: {valor_antigo}')
                                    valores_novos.append(f'{campo}: {valor_novo}')
                                    
                            id_user = Users.objects.get(username=request.user)
                            id_user = id_user.id
                            LogsItens.objects.create(
                                id_user = id_user,
                                nome_user=request.user,
                                nome_objeto=str(venda_nova.slug),
                                acao='Alteração',
                                model = "Vendas_Item",
                                campos_alteracao=', '.join(campos_alteracao),
                                valores_antigos=', '.join(valores_antigos),
                                valores_novos=', '.join(valores_novos)
                            )
                    resultado_venda_controle_filter.venda_finalizada = True #Finalizando a compra
                    resultado_venda_controle_filter.alteracoes_finalizadas = True #Finalizando a compra
                    resultado_venda_controle_filter.troco = troco
                    resultado_venda_controle_filter.valor_realmente_pago = valor_pago
                    resultado_venda_controle_filter.quantidade_parcelas = quantidade_parcelas
                    resultado_venda_controle_filter.save()
                    messages.add_message(request, messages.SUCCESS, 'Venda Confirmada com sucesso')
                    return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))
                else:
                    messages.add_message(request, messages.ERROR, 'Houve um erro na confirmação dessa venda, caso persista contate o administrador')
                    return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug_comunidade}))
            else:
                messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
                return redirect(reverse('home'))
        else:
            messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
            return redirect(reverse('home'))


# ======================== ERRORS ========================
def error_500(request):
    context = {
        'url_atual': 500,
    }
    return render(request, '500.html', context)


def error_404(request, exception):
    context = {
        'url_atual': 404,
    }
    return render(request, '404.html', context)


def error_403(request, exception):
    messages.add_message(request, messages.ERROR, 'Você não tem permissão para isso ou não está logado')
    return redirect(reverse('home'))    


# ======================== EXPORT CSV ========================
@has_permission_decorator('exportar_csv_vendas')
def export_csv_vendas(request, slug):
    opcao = "slug"
    resultado_comunidade = Consultar_Uma_Comunidade(slug, opcao)
    nome_comunidade_id = resultado_comunidade[0]

    Busca = Q(
            Q(nome_comunidade_id=nome_comunidade_id) & Q(venda_finalizada=0)     
    ) 
    venda = Vendas.objects.filter(Busca) #Buscando todas as vendas
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Vendas.xlsx'

    # Cria uma nova planilha
    wb = Workbook()

    # Seleciona a planilha ativa
    ws = wb.active

    if venda:
        if request.user.cargo == "A" or request.user.cargo == "R" or request.user.cargo == "O":
            ws.append(['ID_Venda','Nome_Cliente','Nome_Produto','Quantidade','Preco_Compra','Preco_Venda_Unidade','Preco_Venda_Total','Forma_Venda','Data_Venda','Vendido_Por','Nome_E_Cidade_Da_Comunidade', 'Venda_Finalizada']) # Colunas
        else:
            ws.append(['Nome_Cliente','Nome_Produto','Quantidade','Preco_Venda_Unidade','Preco_Venda_Total','Forma_Venda','Data_Venda','Vendido_Por','Nome_E_Cidade_Da_Comunidade', 'Venda_Finalizada']) # Colunas

        nome_cliente = request.GET.get('nome_cliente_filtro')
        nome_produto = request.GET.get('nome_produto_filtro')
        funcionario = request.GET.get('funcionario')
        dt_start = request.GET.get('dt_start')
        dt_end = request.GET.get('dt_end')

        nome_cliente_novo = unidecode.unidecode(f'{nome_cliente}')

        if nome_cliente or nome_produto or funcionario or dt_start or dt_end:
            if nome_cliente:
                venda = venda.filter(nome_cliente__icontains=nome_cliente_novo)#Filtrando para exportar pelo nome do cliente
            if nome_produto:
                venda = venda.filter(label_vendas_get__icontains=nome_produto)#Filtrando para exportar pelo nome do produto
            if funcionario:
                venda = venda.filter(criado_por__icontains=funcionario)#Filtrando para exportar pelo funcionario
            if dt_start and dt_end:
                venda = venda.filter(dia__range=[dt_start, dt_end])#Filtrando para exportar pelas datas
        else:
            pass
        for itens in venda:
            if itens.venda_finalizada == 1:
                itens.venda_finalizada = "S"
            else:
                itens.venda_finalizada = "N"

            if request.user.cargo == "A" or request.user.cargo == "R" or request.user.cargo == "O":
                row = [itens.id_venda_id, itens.nome_cliente, itens.nome_produto.nome_produto, itens.quantidade, itens.preco_compra, itens.preco_venda, itens.preco_venda_total, itens.forma_venda, itens.data_criacao, itens.criado_por, itens.nome_comunidade.slug, itens.venda_finalizada] #Linhas
            else:
                row = [itens.nome_cliente, itens.nome_produto.nome_produto, itens.quantidade, itens.preco_venda, itens.preco_venda_total, itens.forma_venda, itens.data_criacao, itens.criado_por, itens.nome_comunidade.slug, itens.venda_finalizada] #Linhas
            ws.append(row)
            

        #Inicio formatação da planilha

        # Aplicando filtro em todas as colunas na primeira linha
        ws.auto_filter.ref = ws.dimensions
        
        #Fim formatação da planilha

        # Salva o arquivo   
        wb.save(response)

        return response
    else:
        messages.add_message(request, messages.ERROR, 'Nenhuma venda em andamento foi encontrada para ser exportada')
        return redirect(reverse('consultar_vendas_geral', kwargs={"slug":slug})) 


@has_permission_decorator('exportar_csv_vendas_finalizada')
def export_csv_vendas_finalizadas(request, slug):
    opcao = "slug"
    resultado_comunidade = Consultar_Uma_Comunidade(slug, opcao)
    nome_comunidade_id = resultado_comunidade[0]

    Busca = Q(
            Q(nome_comunidade_id=nome_comunidade_id) & Q(venda_finalizada=1)     
    ) 
    venda = Vendas.objects.filter(Busca) #Buscando todas as vendas
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Vendas_Finalizadas.xlsx'

    # Cria uma nova planilha
    wb = Workbook()

    # Seleciona a planilha ativa
    ws = wb.active

    if venda:
        if request.user.cargo == "A" or request.user.cargo == "R" or request.user.cargo == "O":
            ws.append(['ID_Venda','Nome_Cliente','Nome_Produto','Quantidade','Preco_Compra','Preco_Venda_Unidade','Preco_Venda_Total','Forma_Venda','Data_Venda','Vendido_Por','Nome_E_Cidade_Da_Comunidade', 'Venda_Finalizada']) # Colunas
        else:
            ws.append(['Nome_Cliente','Nome_Produto','Quantidade','Preco_Venda_Unidade','Preco_Venda_Total','Forma_Venda','Data_Venda','Vendido_Por','Nome_E_Cidade_Da_Comunidade', 'Venda_Finalizada']) # Colunas

        nome_cliente = request.GET.get('nome_cliente_filtro')
        nome_produto = request.GET.get('nome_produto_filtro')
        funcionario = request.GET.get('funcionario')
        dt_start = request.GET.get('dt_start')
        dt_end = request.GET.get('dt_end')

        nome_cliente_novo = unidecode.unidecode(f'{nome_cliente}')

        if nome_cliente or nome_produto or funcionario or dt_start or dt_end:
            if nome_cliente:
                venda = venda.filter(nome_cliente__icontains=nome_cliente_novo)#Filtrando para exportar pelo nome do cliente
            if nome_produto:
                venda = venda.filter(label_vendas_get__icontains=nome_produto)#Filtrando para exportar pelo nome do produto
            if funcionario:
                venda = venda.filter(criado_por__icontains=funcionario)#Filtrando para exportar pelo funcionario
            if dt_start and dt_end:
                venda = venda.filter(dia__range=[dt_start, dt_end])#Filtrando para exportar pelas datas
        else:
            pass
        for itens in venda:
            if itens.venda_finalizada == 1:
                itens.venda_finalizada = "S"
            else:
                itens.venda_finalizada = "N"

            if request.user.cargo == "A" or request.user.cargo == "R" or request.user.cargo == "O":
                row = [itens.id_venda_id, itens.nome_cliente, itens.nome_produto.nome_produto, itens.quantidade, itens.preco_compra, itens.preco_venda, itens.preco_venda_total, itens.forma_venda, itens.data_criacao, itens.criado_por, itens.nome_comunidade.slug, itens.venda_finalizada] #Linhas
            else:
                row = [itens.nome_cliente, itens.nome_produto.nome_produto, itens.quantidade, itens.preco_venda, itens.preco_venda_total, itens.forma_venda, itens.data_criacao, itens.criado_por, itens.nome_comunidade.slug, itens.venda_finalizada] #Linhas
            ws.append(row)
            

        #Inicio formatação da planilha

        # Aplicando filtro em todas as colunas na primeira linha
        ws.auto_filter.ref = ws.dimensions
        
        #Fim formatação da planilha

        # Salva o arquivo   
        wb.save(response)

        return response
    else:
        messages.add_message(request, messages.ERROR, 'Nenhuma venda finalizada foi encontrada para ser exportada')
        return redirect(reverse('vendas_finalizadas', kwargs={"slug":slug})) 


#Função para exportar produtos
@has_permission_decorator('exportar_csv_p')
def export_csv_produto(request, slug):
    opcao = "slug"
    resultado = Consultar_Uma_Comunidade(slug, opcao)

    if resultado[0] != 0:
        id_comunidade_comparar_usuario, id_comunidade_usuario = Bloqueio_Acesso_Demais_Comunidades(request, resultado[0])
        if id_comunidade_comparar_usuario == id_comunidade_usuario:
            pass
        else:
            messages.add_message(request, messages.ERROR, 'Não foram encontradas dados referente à comunidade escolhida')
            return redirect(reverse('home'))
    else:
        messages.add_message(request, messages.ERROR, 'Essa URL que você tentou acessar não foi encontrada')
        return redirect(reverse('export_entrada_produtos', kwargs={"slug":slug})) 

    Busca = Q(
            Q(nome_e_cidade_comunidade=resultado[1])     
    ) 

    p_excel = P_Excel.objects.filter(Busca) #Buscando todas os produtos do dia à exportar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Estoque.xlsx'
    
    # Cria uma nova planilha
    wb = Workbook()

    # Seleciona a planilha ativa
    ws = wb.active
    
    if p_excel:
        if request.user.cargo == "A" or request.user.cargo == "R" or request.user.cargo == "O":
            ws.append(['Acao','Nome_Produto','Quantidade','Data Criação','Criado Por','Última Alteração','Alterado Por','Nome_E_Cidade_Comunidade']) # Colunas
        else:
            ws.append(['Acao','Nome_Produto','Quantidade','Data','Nome_E_Cidade_Comunidade']) # Colunas

        nome_produto = request.GET.get('nome_produto')
        acao = request.GET.get('acao')
        dia = request.GET.get('dia')

        if nome_produto or acao or dia:
            if nome_produto:
                p_excel = p_excel.filter(nome_produto__icontains=nome_produto)#Filtrando para exportar pelo nome do produto
            if acao:
                p_excel = p_excel.filter(acao=acao)#Filtrando para exportar pela ação
            if dia:
                p_excel = p_excel.filter(dia=dia)#Filtrando para exportar pela data
        else:
            pass

        for itens in p_excel:
            if request.user.cargo == "A" or request.user.cargo == "R" or request.user.cargo == "O":
                row = [itens.acao, itens.nome_produto, itens.quantidade, itens.data, itens.nome_user, itens.ultima_alteracao, itens.alterado_por, itens.nome_e_cidade_comunidade]
            else:
                row = [itens.acao, itens.nome_produto, itens.quantidade, itens.data, itens.nome_e_cidade_comunidade]
            ws.append(row)

        #Inicio formatação da planilha

        # Define estilos para as células
        blue_font = Font(color="0000FF")  # Azul
        red_font = Font(color="FF0000")    # Vermelho
        bold_font = Font(bold=True)

        # Aplica os estilos aos títulos
        for cell in ws[1]:  # A primeira linha contém os títulos
            cell.font = bold_font

        # Aplica os estilos condicionais à coluna 'Acao' (min_col=1) e (max_col=1) significa que ele só vai procurar na primeira coluna
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # Ignora a primeira linha (títulos)
            for cell in row:
                if cell.value == 'Entrada':
                    cell.font = blue_font
                elif cell.value == 'Saída':
                    cell.font = red_font

        # Aplicando filtro em todas as colunas na primeira linha
        ws.auto_filter.ref = ws.dimensions

        #Fim formatação da planilha

        # Salva o arquivo   
        wb.save(response)

        return response
    else:
        messages.add_message(request, messages.ERROR, 'Nenhuma ação de produto foi encontrado para ser exportada')
        return redirect(reverse('add_produto', kwargs={"slug":slug})) 


#Função para a tela de export_entrada_produtos
@has_permission_decorator('exportar_csv_p')
def export_entrada_produtos(request, slug):
    if request.method == "GET":
        nome_produto = request.GET.get('nome_produto')
        dia = request.GET.get('dia')
        acao = request.GET.get('acao')
        
        paginacao = P_Excel.objects.filter(nome_e_cidade_comunidade=slug)
        if paginacao:
            validacao, page = Get_Paginacao(request, nome_produto, dia, acao, slug, paginacao)        
            if validacao:
                return validacao
            
            url_atual = Capturar_Url_Atual_Sem_O_Final(request)
            context = {
                'page': page,
                'slug': slug,
                'url_atual': url_atual,
            }
            return render(request, 'export_entrada_produtos.html', context)
        else:
            messages.add_message(request, messages.ERROR, f'Não há registro de Logs à serem exportados')
            return redirect(reverse('add_produto', kwargs={"slug":slug})) 


# Pegando a última palavra de qualquer string
def palavra_final(string):
    padrao = r'\b(\w+)\s*$'
    resultado = re.search(padrao, string)
    if resultado:
        return resultado.group(1)
    else:
        return ""


# Pegando a primeira palavra de qualquer string
def primeira_palavra(string):
    padrao = r'^\s*([\w()\-]+)'
    resultado = re.search(padrao, string)
    if resultado:
        return resultado.group(1)
    else:
        return ""


# Pegando a segunda palavra de qualquer string após um espaço
def segunda_palavra(string):
    padrao = r'\s(\w+)'
    resultado = re.search(padrao, string)
    if resultado:
        return resultado.group(1)
    else:
        return ""